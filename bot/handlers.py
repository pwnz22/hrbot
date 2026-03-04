from aiogram import Dispatcher
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.filters.callback_data import CallbackData
from sqlalchemy import desc, select
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from shared.database.database import AsyncSessionLocal
from shared.models.vacancy import Application, Vacancy
from shared.models.user import TelegramUser, RoleEnum
from shared.services.resume_summary_service import ResumeSummaryService
from bot.middleware import moderator_or_admin, admin_only

# Словарь для хранения ID сообщений с файлами резюме для каждого пользователя
user_resume_messages = {}

# Словарь для хранения состояния ввода описания обработки
user_description_states = {}  # {user_id: {"application_id": int, "action": "add"|"edit"}}

class VacancyCallback(CallbackData, prefix="vacancy"):
    vacancy_id: int

class ApplicationCallback(CallbackData, prefix="application"):
    application_id: int
    source: str = "recent"  # "recent", "unprocessed" или "vacancy"

class ProcessCallback(CallbackData, prefix="process"):
    application_id: int
    action: str  # "mark_processed" или "mark_unprocessed"

class BackCallback(CallbackData, prefix="back"):
    to: str  # "vacancies", "applications" или "unprocessed"
    vacancy_id: int = 0  # Для возврата к приложениям конкретной вакансии

class DeleteCallback(CallbackData, prefix="delete"):
    application_id: int
    action: str  # "confirm" или "cancel"
    source: str = "recent"  # Откуда пришел пользователь

class SummaryCallback(CallbackData, prefix="summary"):
    application_id: int
    action: str  # "generate", "show"

class QuestionsCallback(CallbackData, prefix="questions"):
    application_id: int
    action: str  # "generate", "show"

class ResumeCallback(CallbackData, prefix="resume"):
    application_id: int
    action: str  # "download"

class AccountCallback(CallbackData, prefix="account"):
    account_id: str

class AccountToggleCallback(CallbackData, prefix="account_toggle"):
    account_id: str
    action: str  # "enable" или "disable"

class UserCallback(CallbackData, prefix="user"):
    user_id: int

class UserRoleCallback(CallbackData, prefix="user_role"):
    user_id: int
    role: str  # "user", "moderator", "admin"

class AccountLinkCallback(CallbackData, prefix="account_link"):
    account_id: str
    action: str  # "show_users", "link", "unlink"
    user_id: int = 0

class AccountDeleteCallback(CallbackData, prefix="account_delete"):
    account_id: str
    action: str  # "confirm", "execute"

class AccountAuthCallback(CallbackData, prefix="account_auth"):
    account_id: str

class DescriptionCallback(CallbackData, prefix="description"):
    application_id: int
    action: str  # "view", "edit"
    source: str = "recent"

class ExportCallback(CallbackData, prefix="export"):
    filter_type: str  # "all" или "unprocessed"

async def delete_message_after_delay(message, delay_seconds):
    """Удаляет сообщение через указанное количество секунд"""
    import asyncio
    await asyncio.sleep(delay_seconds)
    try:
        await message.delete()
    except Exception:
        pass  # Игнорируем ошибки если сообщение уже удалено

def clean_html_tags(text):
    """Удаляет HTML теги и эмодзи из текста для Excel"""
    if not text:
        return ""

    import re
    # Удаляем HTML теги
    text = re.sub(r'<[^>]+>', '', text)
    # Удаляем эмодзи (базовые)
    text = re.sub(r'[📋👤📧📱🔗📝🛠⏰🎓⚠️📊🤖📄]', '', text)
    # Заменяем множественные пробелы на одинарные
    text = re.sub(r'\s+', ' ', text)
    # Убираем пробелы в начале и конце
    text = text.strip()
    return text

def setup_handlers(dp: Dispatcher):

    @dp.message(CommandStart())
    async def command_start_handler(message: Message, user: TelegramUser) -> None:
        from aiogram.types import ReplyKeyboardMarkup, KeyboardButton

        # Формируем клавиатуру и текст в зависимости от роли
        if user.is_admin:
            keyboard = ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📋 Последние"), KeyboardButton(text="⏳ Необработанные")],
                    [KeyboardButton(text="🔄 Парсинг"), KeyboardButton(text="📊 Статистика")],
                    [KeyboardButton(text="📧 Аккаунты"), KeyboardButton(text="➕ Добавить аккаунт")],
                    [KeyboardButton(text="📥 Экспорт"), KeyboardButton(text="👥 Пользователи")]
                ],
                resize_keyboard=True
            )

            help_text = (
                "Привет! Я HR-бот для обработки откликов на вакансии из Gmail.\n\n"
                f"👤 Ваша роль: <b>Администратор</b>\n\n"
                "<b>Команды:</b>\n"
                "/start - Это сообщение\n"
                "/stats - Статистика по откликам\n"
                "/recent - Последние отклики\n"
                "/unprocessed - Все необработанные отклики\n"
                "/parse - Парсить новые письма\n"
                "/export - Экспорт откликов в Excel\n"
                "/accounts - Управление Gmail аккаунтами\n"
                "/add_account - Добавить новый Gmail аккаунт\n"
                "/users - Управление пользователями"
            )

        elif user.is_moderator:
            keyboard = ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📋 Последние"), KeyboardButton(text="⏳ Необработанные")],
                    [KeyboardButton(text="🔄 Парсинг"), KeyboardButton(text="📊 Статистика")],
                    [KeyboardButton(text="📥 Экспорт")]
                ],
                resize_keyboard=True
            )

            help_text = (
                "Привет! Я HR-бот для обработки откликов на вакансии из Gmail.\n\n"
                f"👤 Ваша роль: <b>Модератор</b>\n\n"
                "<b>Команды:</b>\n"
                "/start - Это сообщение\n"
                "/stats - Статистика по откликам\n"
                "/recent - Последние отклики\n"
                "/unprocessed - Все необработанные отклики\n"
                "/parse - Парсить новые письма\n"
                "/export - Экспорт откликов в Excel"
            )

        else:  # USER
            keyboard = ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="🏠 Главная")]
                ],
                resize_keyboard=True
            )

            help_text = (
                "Привет! Я HR-бот для обработки откликов на вакансии из Gmail.\n\n"
                f"👤 Ваша роль: <b>Пользователь</b>\n\n"
                "У вас нет прав для работы с ботом.\n"
                "Обратитесь к администратору для получения доступа."
            )

        await message.answer(help_text, reply_markup=keyboard, parse_mode="HTML")

    @dp.message(Command("stats"))
    @moderator_or_admin
    async def stats_handler(message: Message, user: TelegramUser) -> None:
        async with AsyncSessionLocal() as session:
            from sqlalchemy import func, case
            from shared.models.gmail_account import GmailAccount

            # Получаем аккаунты в зависимости от роли
            if user.is_admin:
                accounts_stmt = select(GmailAccount).where(GmailAccount.enabled == True)
            else:
                # Модератор видит только свои привязанные аккаунты
                accounts_stmt = select(GmailAccount).where(
                    GmailAccount.enabled == True,
                    GmailAccount.user_id == user.id
                )

            accounts_result = await session.execute(accounts_stmt)
            accounts = accounts_result.scalars().all()

            # Логирование для отладки
            print(f"User role: {user.role}, User ID: {user.id}, Is Admin: {user.is_admin}")
            print(f"Found {len(accounts)} accounts for user")

            if not accounts:
                await message.answer("❌ Нет доступных аккаунтов для статистики")
                return

            text = "📊 <b>Статистика по аккаунтам:</b>\n\n"

            total_vacancies = 0
            total_applications = 0
            total_processed = 0
            total_unprocessed = 0

            for account in accounts:
                # Статистика по вакансиям для этого аккаунта
                vacancies_count_stmt = select(func.count(Vacancy.id)).where(
                    Vacancy.gmail_account_id == account.id
                )
                vacancies_count = await session.scalar(vacancies_count_stmt)

                # Статистика по откликам для этого аккаунта
                applications_stmt = select(
                    func.count(Application.id).label('total'),
                    func.sum(case((Application.is_processed == True, 1), else_=0)).label('processed'),
                    func.sum(case((Application.is_processed == False, 1), else_=0)).label('unprocessed')
                ).select_from(Application).join(
                    Vacancy, Application.vacancy_id == Vacancy.id
                ).where(
                    Vacancy.gmail_account_id == account.id,
                    Application.deleted_at.is_(None)
                )

                result = await session.execute(applications_stmt)
                stats = result.one()

                apps_total = stats.total or 0
                apps_processed = stats.processed or 0
                apps_unprocessed = stats.unprocessed or 0

                total_vacancies += vacancies_count
                total_applications += apps_total
                total_processed += apps_processed
                total_unprocessed += apps_unprocessed

                text += f"📧 <b>{account.name}</b>\n"
                text += f"   📋 Вакансий: {vacancies_count}\n"
                text += f"   👥 Откликов: {apps_total}\n"
                text += f"   ✅ Обработано: {apps_processed}\n"
                text += f"   ❌ Не обработано: {apps_unprocessed}\n\n"

            # Общая статистика
            text += f"━━━━━━━━━━━━━━━━━━━━\n"
            text += f"📊 <b>Итого:</b>\n"
            text += f"📧 Аккаунтов: <b>{len(accounts)}</b>\n"
            text += f"📋 Вакансий: <b>{total_vacancies}</b>\n"
            text += f"👥 Откликов: <b>{total_applications}</b>\n"
            text += f"✅ Обработано: <b>{total_processed}</b>\n"
            text += f"❌ Не обработано: <b>{total_unprocessed}</b>"

            await message.answer(text, parse_mode="HTML")

    @dp.message(Command("recent"))
    @moderator_or_admin
    async def recent_handler(message: Message, user: TelegramUser) -> None:
        async with AsyncSessionLocal() as session:
            # Получаем список вакансий с фильтрацией по пользователю
            if user.is_admin:
                # Админ видит все вакансии (включая те, где gmail_account_id = NULL)
                stmt = select(Vacancy).order_by(desc(Vacancy.created_at))
            else:
                # Модератор видит только вакансии привязанных к нему аккаунтов
                from shared.models.gmail_account import GmailAccount
                stmt = select(Vacancy).join(
                    GmailAccount, Vacancy.gmail_account_id == GmailAccount.id, isouter=False
                ).where(
                    GmailAccount.user_id == user.id
                ).order_by(desc(Vacancy.created_at))

            result = await session.execute(stmt)
            vacancies = result.scalars().all()

            if not vacancies:
                await message.answer("Пока нет вакансий")
                return

            # Создаем клавиатуру с вакансиями
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for vacancy in vacancies:
                # Считаем количество откликов для каждой вакансии (исключаем удаленные)
                count_stmt = select(Application).where(Application.vacancy_id == vacancy.id, Application.deleted_at.is_(None))
                count_result = await session.execute(count_stmt)
                applications_count = len(count_result.scalars().all())

                button_text = f"{vacancy.title} ({applications_count} откликов)"
                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=VacancyCallback(vacancy_id=vacancy.id).pack()
                )
                keyboard.inline_keyboard.append([button])

            await message.answer("📋 Выберите вакансию для просмотра откликов:", reply_markup=keyboard)

    @dp.message(Command("parse"))
    @moderator_or_admin
    async def parse_handler(message: Message, user: TelegramUser) -> None:
        status_msg = await message.answer("🔄 Начинаю парсинг новых писем...")

        try:
            import json
            import os
            from bot.gmail_parser import GmailParser
            from shared.models.gmail_account import GmailAccount

            # Получаем список аккаунтов для парсинга в зависимости от роли
            async with AsyncSessionLocal() as session:
                if user.is_admin:
                    # Админ парсит все активные аккаунты
                    stmt = select(GmailAccount).where(GmailAccount.enabled == True)
                else:
                    # Модератор парсит только свои привязанные аккаунты
                    stmt = select(GmailAccount).where(
                        GmailAccount.enabled == True,
                        GmailAccount.user_id == user.id
                    )

                result = await session.execute(stmt)
                accounts_from_db = result.scalars().all()

            parsers = []

            # Создаем парсеры для найденных аккаунтов
            for account in accounts_from_db:
                if os.path.exists(account.credentials_path) and os.path.exists(account.token_path):
                    parser = GmailParser(
                        account_id=account.account_id,
                        credentials_path=account.credentials_path,
                        token_path=account.token_path,
                        sender_email=account.sender_email
                    )
                    parsers.append(parser)

            if not parsers:
                await status_msg.edit_text("❌ Нет доступных аккаунтов для парсинга")
                return

            # Парсим новые письма из всех аккаунтов
            total_parsed = 0
            all_new_vacancies = []

            for parser in parsers:
                result = await parser.parse_new_emails()
                total_parsed += result["parsed_count"]

                if result["new_vacancies"]:
                    all_new_vacancies.extend(result["new_vacancies"])

            if total_parsed > 0:
                text = f"✅ Парсинг завершен!\nОбработано новых откликов: <b>{total_parsed}</b>"

                if all_new_vacancies:
                    unique_vacancies = list(set(all_new_vacancies))
                    text += f"\n\n<b>Новые вакансии ({len(unique_vacancies)}):</b>"
                    for vacancy in unique_vacancies:
                        text += f"\n• {vacancy}"

                await status_msg.edit_text(text, parse_mode="HTML")
            else:
                await status_msg.edit_text("📭 Новых писем не найдено")

            # Удаляем сообщения через 2 секунды
            import asyncio
            asyncio.create_task(delete_message_after_delay(status_msg, 2))
            asyncio.create_task(delete_message_after_delay(message, 2))

        except Exception as e:
            await status_msg.edit_text(f"❌ Ошибка при парсинге: {str(e)}")
            # Удаляем сообщения с ошибкой через 2 секунды
            import asyncio
            asyncio.create_task(delete_message_after_delay(status_msg, 2))
            asyncio.create_task(delete_message_after_delay(message, 2))

    @dp.message(Command("unprocessed"))
    @moderator_or_admin
    async def unprocessed_handler(message: Message, user: TelegramUser) -> None:
        async with AsyncSessionLocal() as session:
            # Получаем все необработанные отклики с фильтрацией по пользователю
            from sqlalchemy.orm import selectinload

            if user.is_admin:
                # Админ видит все необработанные отклики
                stmt = select(Application).options(selectinload(Application.vacancy)).where(
                    Application.is_processed == False,
                    Application.deleted_at.is_(None)
                ).order_by(desc(Application.created_at))
            else:
                # Модератор видит только отклики из привязанных к нему аккаунтов
                from shared.models.gmail_account import GmailAccount
                stmt = select(Application).options(
                    selectinload(Application.vacancy).selectinload(Vacancy.gmail_account)
                ).join(
                    Vacancy, Application.vacancy_id == Vacancy.id, isouter=False
                ).join(
                    GmailAccount, Vacancy.gmail_account_id == GmailAccount.id, isouter=False
                ).where(
                    Application.is_processed == False,
                    Application.deleted_at.is_(None),
                    GmailAccount.user_id == user.id
                ).order_by(desc(Application.created_at))

            result = await session.execute(stmt)
            unprocessed_applications = result.scalars().all()

            if not unprocessed_applications:
                await message.answer("✅ Все отклики обработаны!")
                return

            # Создаем клавиатуру с необработанными откликами
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for app in unprocessed_applications:
                # Формируем текст кнопки: {имя} - {телефон} - {email}
                button_text = app.name
                if app.phone:
                    button_text += f" - {app.phone}"
                if app.email:
                    button_text += f" - {app.email}"

                # Добавляем эмодзи для необработанных
                button_text = f"❌ {button_text}"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=ApplicationCallback(application_id=app.id, source="unprocessed").pack()
                )
                keyboard.inline_keyboard.append([button])

            text = f"❌ <b>Необработанные отклики ({len(unprocessed_applications)}):</b>\n\n"
            text += "Выберите отклик для просмотра и обработки:"

            await message.answer(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(VacancyCallback.filter())
    async def vacancy_applications_handler(query: CallbackQuery, callback_data: VacancyCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав для просмотра откликов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Получаем вакансию
            vacancy_stmt = select(Vacancy).where(Vacancy.id == callback_data.vacancy_id)
            vacancy_result = await session.execute(vacancy_stmt)
            vacancy = vacancy_result.scalar_one_or_none()

            if not vacancy:
                await query.message.edit_text("Вакансия не найдена")
                return

            # Получаем отклики на эту вакансию (исключаем удаленные)
            apps_stmt = select(Application).where(Application.vacancy_id == callback_data.vacancy_id, Application.deleted_at.is_(None)).order_by(desc(Application.created_at))
            apps_result = await session.execute(apps_stmt)
            applications = apps_result.scalars().all()

            if not applications:
                await query.message.edit_text(f"📋 Вакансия: {vacancy.title}\n\nОткликов пока нет")
                return

            # Создаем клавиатуру с откликами
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for app in applications:
                # Формируем текст кнопки: {имя} - {телефон} - {email}
                button_text = app.name
                if app.phone:
                    button_text += f" - {app.phone}"
                if app.email:
                    button_text += f" - {app.email}"

                # Добавляем статус обработки
                status = "✅" if app.is_processed else "❌"
                button_text = f"{status} {button_text}"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=ApplicationCallback(application_id=app.id, source="vacancy").pack()
                )
                keyboard.inline_keyboard.append([button])

            # Добавляем кнопку "Назад к вакансиям"
            back_button = InlineKeyboardButton(
                text="⬅️ Назад к вакансиям",
                callback_data=BackCallback(to="vacancies").pack()
            )
            keyboard.inline_keyboard.append([back_button])

            text = f"📋 Вакансия: <b>{vacancy.title}</b>\n"
            text += f"📊 Всего откликов: <b>{len(applications)}</b>\n\n"
            text += "Выберите отклик для просмотра подробностей:"

            await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(ApplicationCallback.filter())
    async def application_details_handler(query: CallbackQuery, callback_data: ApplicationCallback, user: TelegramUser) -> None:
        from bot.utils.formatters import format_application_details

        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав для просмотра откликов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Получаем отклик
            app_stmt = select(Application).where(Application.id == callback_data.application_id)
            app_result = await session.execute(app_stmt)
            application = app_result.scalar_one_or_none()

            if not application:
                await query.message.edit_text("Отклик не найден")
                return

            # Получаем вакансию
            vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
            vacancy_result = await session.execute(vacancy_stmt)
            vacancy = vacancy_result.scalar_one_or_none()

            # Формируем детальную информацию
            text = format_application_details(application, vacancy, include_description=False)


            # Создаем кнопки для изменения статуса обработки и навигации
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])
            if application.is_processed:
                process_button = InlineKeyboardButton(
                    text="❌ Отменить обработку",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_unprocessed").pack()
                )
            else:
                process_button = InlineKeyboardButton(
                    text="✅ Отметить как обработанный",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_processed").pack()
                )
            # Добавляем кнопку удаления
            delete_button = InlineKeyboardButton(
                text="🗑️ Удалить отклик",
                callback_data=DeleteCallback(application_id=application.id, action="confirm", source=callback_data.source).pack()
            )
            keyboard.inline_keyboard.append([process_button, delete_button])

            # Добавляем кнопку просмотра описания только если описание есть
            if application.processing_description:
                description_button = InlineKeyboardButton(
                    text="📝 Посмотреть описание",
                    callback_data=DescriptionCallback(application_id=application.id, action="view", source=callback_data.source).pack()
                )
                keyboard.inline_keyboard.append([description_button])

            # Добавляем кнопки для работы с резюме если есть файл
            if (application.file_path or application.attachment_filename) and (
                application.attachment_filename and
                (application.attachment_filename.lower().endswith('.pdf') or application.attachment_filename.lower().endswith('.docx'))
            ):
                # Кнопка для работы с summary
                if application.summary:
                    # Если summary уже есть - показываем кнопку для отправки
                    summary_button = InlineKeyboardButton(
                        text="📊 Показать анализ резюме",
                        callback_data=SummaryCallback(application_id=application.id, action="show").pack()
                    )
                else:
                    # Если summary нет - показываем кнопку для генерации
                    summary_button = InlineKeyboardButton(
                        text="🤖 Сгенерировать анализ резюме",
                        callback_data=SummaryCallback(application_id=application.id, action="generate").pack()
                    )
                keyboard.inline_keyboard.append([summary_button])

                # Кнопка получения файла резюме
                resume_button = InlineKeyboardButton(
                    text="📄 Получить файл резюме",
                    callback_data=ResumeCallback(application_id=application.id, action="download").pack()
                )
                keyboard.inline_keyboard.append([resume_button])

            # Добавляем кнопку "Назад" в зависимости от источника
            if callback_data.source == "unprocessed":
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад к необработанным",
                    callback_data=BackCallback(to="unprocessed").pack()
                )
            elif callback_data.source == "vacancy":
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад к откликам",
                    callback_data=BackCallback(to="applications", vacancy_id=application.vacancy_id).pack()
                )
            else:  # source == "recent" или другое
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад к вакансиям",
                    callback_data=BackCallback(to="vacancies").pack()
                )
            keyboard.inline_keyboard.append([back_button])

            # Отправляем сообщение с информацией
            await query.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)

    @dp.callback_query(ProcessCallback.filter())
    async def process_status_handler(query: CallbackQuery, callback_data: ProcessCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('change_status'):
            await query.answer("❌ У вас нет прав для изменения статуса", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Получаем отклик
            app_stmt = select(Application).where(Application.id == callback_data.application_id)
            app_result = await session.execute(app_stmt)
            application = app_result.scalar_one_or_none()

            if not application:
                await query.message.edit_text("❌ Отклик не найден")
                return

            # Изменяем статус обработки
            if callback_data.action == "mark_processed":
                # Запрашиваем описание вместо немедленной обработки
                user_description_states[query.from_user.id] = {
                    "application_id": application.id,
                    "action": "add"
                }

                prompt_text = (
                    "📝 <b>Описание обработки</b>\n\n"
                    f"Вы отмечаете отклик кандидата <b>{application.name}</b> как обработанный.\n\n"
                    "Пожалуйста, введите описание обработки (обязательно):\n"
                    "• Результат собеседования\n"
                    "• Комментарии\n"
                    "• Дальнейшие действия\n\n"
                    "💡 Отправьте /cancel для отмены"
                )
                await query.message.edit_text(prompt_text, parse_mode="HTML")
                return

            elif callback_data.action == "mark_unprocessed":
                application.is_processed = False
                # Сохраняем описание при отмене обработки
                status_message = "❌ Обработка отклика отменена"

            # Сохраняем изменения
            await session.commit()

            # Получаем вакансию для обновления информации
            vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
            vacancy_result = await session.execute(vacancy_stmt)
            vacancy = vacancy_result.scalar_one_or_none()

            # Обновляем сообщение с новой информацией
            from bot.utils.formatters import format_application_details
            text = format_application_details(application, vacancy, include_description=False)


            # Создаем обновленные кнопки
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])
            if application.is_processed:
                process_button = InlineKeyboardButton(
                    text="❌ Отменить обработку",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_unprocessed").pack()
                )
            else:
                process_button = InlineKeyboardButton(
                    text="✅ Отметить как обработанный",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_processed").pack()
                )
            # Добавляем кнопку удаления
            delete_button = InlineKeyboardButton(
                text="🗑️ Удалить отклик",
                callback_data=DeleteCallback(application_id=application.id, action="confirm", source="vacancy").pack()
            )
            keyboard.inline_keyboard.append([process_button, delete_button])

            # Добавляем кнопку просмотра описания только если описание есть
            if application.processing_description:
                description_button = InlineKeyboardButton(
                    text="📝 Посмотреть описание",
                    callback_data=DescriptionCallback(application_id=application.id, action="view", source="vacancy").pack()
                )
                keyboard.inline_keyboard.append([description_button])

            # Получаем source из исходного callback (нужно передать через ProcessCallback)
            # Пока используем applications как fallback
            back_button = InlineKeyboardButton(
                text="⬅️ Назад к откликам",
                callback_data=BackCallback(to="applications", vacancy_id=application.vacancy_id).pack()
            )
            keyboard.inline_keyboard.append([back_button])

            # Обновляем сообщение
            await query.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)

            # Отправляем уведомление об изменении статуса и удаляем через секунду
            status_msg = await query.message.answer(status_message)

            # Удаляем сообщение через 1 секунду
            import asyncio
            asyncio.create_task(delete_message_after_delay(status_msg, 1))

    @dp.message(lambda message: message.from_user.id in user_description_states and message.text and not message.text.startswith('/'))
    async def handle_description_input(message: Message, user: TelegramUser) -> None:
        """Обрабатывает ввод описания обработки отклика"""
        from bot.utils.formatters import format_application_details

        if not user.has_permission('change_status'):
            await message.answer("❌ У вас нет прав для изменения статуса")
            if message.from_user.id in user_description_states:
                del user_description_states[message.from_user.id]
            return

        state_data = user_description_states.get(message.from_user.id)
        if not state_data:
            return

        application_id = state_data["application_id"]
        action = state_data["action"]
        description_text = message.text.strip()

        # Валидация
        if len(description_text) < 3:
            await message.answer(
                "⚠️ Описание слишком короткое. Введите хотя бы 3 символа.\n"
                "Отправьте /cancel для отмены."
            )
            return

        if len(description_text) > 4000:
            await message.answer(
                "⚠️ Описание слишком длинное (максимум 4000 символов).\n"
                "Отправьте /cancel для отмены."
            )
            return

        async with AsyncSessionLocal() as session:
            # Получаем отклик
            app_stmt = select(Application).where(Application.id == application_id)
            app_result = await session.execute(app_stmt)
            application = app_result.scalar_one_or_none()

            if not application:
                del user_description_states[message.from_user.id]
                await message.answer("❌ Отклик не найден")
                return

            # Сохраняем описание и обновляем статус если нужно
            application.processing_description = description_text
            if action == "add":
                application.is_processed = True

            await session.commit()

            # Получаем вакансию
            vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
            vacancy_result = await session.execute(vacancy_stmt)
            vacancy = vacancy_result.scalar_one_or_none()

            # Очищаем состояние
            del user_description_states[message.from_user.id]

            # Формируем ответ
            text = format_application_details(application, vacancy, include_description=False)

            # Создаём клавиатуру
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            # Кнопки обработки и удаления
            if application.is_processed:
                process_button = InlineKeyboardButton(
                    text="❌ Отменить обработку",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_unprocessed").pack()
                )
            else:
                process_button = InlineKeyboardButton(
                    text="✅ Отметить как обработанный",
                    callback_data=ProcessCallback(application_id=application.id, action="mark_processed").pack()
                )

            delete_button = InlineKeyboardButton(
                text="🗑️ Удалить отклик",
                callback_data=DeleteCallback(application_id=application.id, action="confirm", source="recent").pack()
            )
            keyboard.inline_keyboard.append([process_button, delete_button])

            # Кнопка просмотра описания только если описание есть
            if application.processing_description:
                description_button = InlineKeyboardButton(
                    text="📝 Посмотреть описание",
                    callback_data=DescriptionCallback(application_id=application.id, action="view", source="recent").pack()
                )
                keyboard.inline_keyboard.append([description_button])

            # Кнопки для резюме
            if (application.file_path or application.attachment_filename) and (
                application.attachment_filename and
                (application.attachment_filename.lower().endswith('.pdf') or application.attachment_filename.lower().endswith('.docx'))
            ):
                if application.summary:
                    summary_button = InlineKeyboardButton(
                        text="📊 Показать анализ резюме",
                        callback_data=SummaryCallback(application_id=application.id, action="show").pack()
                    )
                else:
                    summary_button = InlineKeyboardButton(
                        text="🤖 Сгенерировать анализ резюме",
                        callback_data=SummaryCallback(application_id=application.id, action="generate").pack()
                    )
                keyboard.inline_keyboard.append([summary_button])

                resume_button = InlineKeyboardButton(
                    text="📄 Получить файл резюме",
                    callback_data=ResumeCallback(application_id=application.id, action="download").pack()
                )
                keyboard.inline_keyboard.append([resume_button])

            # Кнопка назад
            back_button = InlineKeyboardButton(
                text="⬅️ Назад к вакансиям",
                callback_data=BackCallback(to="vacancies").pack()
            )
            keyboard.inline_keyboard.append([back_button])

            # Отправляем сообщение
            await message.answer(text, parse_mode="HTML", reply_markup=keyboard)

            # Уведомление об успехе
            success_text = "✅ Описание сохранено!" if action == "edit" else "✅ Отклик отмечен как обработанный с описанием!"
            status_msg = await message.answer(success_text)

            # Удаляем сообщения через 2 секунды
            import asyncio
            asyncio.create_task(delete_message_after_delay(status_msg, 2))
            asyncio.create_task(delete_message_after_delay(message, 2))

    @dp.callback_query(DescriptionCallback.filter())
    async def description_handler(query: CallbackQuery, callback_data: DescriptionCallback, user: TelegramUser) -> None:
        """Обрабатывает просмотр и редактирование описания обработки"""
        from bot.utils.formatters import format_application_details

        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав для просмотра откликов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Получаем отклик
            app_stmt = select(Application).where(Application.id == callback_data.application_id)
            app_result = await session.execute(app_stmt)
            application = app_result.scalar_one_or_none()

            if not application:
                await query.message.edit_text("❌ Отклик не найден")
                return

            # Получаем вакансию
            vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
            vacancy_result = await session.execute(vacancy_stmt)
            vacancy = vacancy_result.scalar_one_or_none()

            if callback_data.action == "view":
                # Показываем полную информацию с описанием
                text = "📝 <b>Описание обработки отклика</b>\n\n"
                text += format_application_details(application, vacancy, include_description=True)

                # Создаём клавиатуру
                keyboard = InlineKeyboardMarkup(inline_keyboard=[])

                # Кнопка редактирования
                if user.has_permission('change_status'):
                    edit_button = InlineKeyboardButton(
                        text="✏️ Редактировать описание",
                        callback_data=DescriptionCallback(
                            application_id=application.id,
                            action="edit",
                            source=callback_data.source
                        ).pack()
                    )
                    keyboard.inline_keyboard.append([edit_button])

                # Кнопка назад к отклику
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад к отклику",
                    callback_data=ApplicationCallback(
                        application_id=application.id,
                        source=callback_data.source
                    ).pack()
                )
                keyboard.inline_keyboard.append([back_button])

                await query.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)

            elif callback_data.action == "edit":
                # Проверяем права
                if not user.has_permission('change_status'):
                    await query.answer("❌ У вас нет прав для редактирования", show_alert=True)
                    return

                # Устанавливаем состояние редактирования
                user_description_states[query.from_user.id] = {
                    "application_id": application.id,
                    "action": "edit"
                }

                current_desc = application.processing_description or "Отсутствует"

                prompt_text = (
                    "✏️ <b>Редактирование описания обработки</b>\n\n"
                    f"Кандидат: <b>{application.name}</b>\n\n"
                    f"<b>Текущее описание:</b>\n{current_desc}\n\n"
                    "Отправьте новое описание следующим сообщением или /cancel для отмены:"
                )
                await query.message.edit_text(prompt_text, parse_mode="HTML")

    @dp.callback_query(DeleteCallback.filter())
    async def delete_handler(query: CallbackQuery, callback_data: DeleteCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('change_status'):
            await query.answer("❌ У вас нет прав для удаления откликов", show_alert=True)
            return

        if callback_data.action == "confirm":
            # Показываем подтверждение удаления
            async with AsyncSessionLocal() as session:
                app_stmt = select(Application).where(Application.id == callback_data.application_id)
                app_result = await session.execute(app_stmt)
                application = app_result.scalar_one_or_none()

                if not application:
                    await query.message.edit_text("❌ Отклик не найден")
                    return

                # Создаем клавиатуру подтверждения
                keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Да, удалить",
                            callback_data=DeleteCallback(application_id=application.id, action="execute", source=callback_data.source).pack()
                        ),
                        InlineKeyboardButton(
                            text="❌ Отмена",
                            callback_data=DeleteCallback(application_id=application.id, action="cancel", source=callback_data.source).pack()
                        )
                    ]
                ])

                text = f"⚠️ <b>Подтверждение удаления</b>\n\n"
                text += f"Вы действительно хотите удалить отклик?\n\n"
                text += f"👤 <b>{application.name}</b>\n"
                if application.email:
                    text += f"📧 {application.email}\n"
                if application.phone:
                    text += f"📱 {application.phone}\n"
                text += f"\n❗ <b>Это действие нельзя отменить!</b>"

                await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

        elif callback_data.action == "execute":
            # Удаляем сообщение с файлом резюме если оно есть
            user_id = query.from_user.id
            if user_id in user_resume_messages:
                try:
                    await query.bot.delete_message(chat_id=query.message.chat.id, message_id=user_resume_messages[user_id])
                except Exception:
                    pass  # Игнорируем ошибки если сообщение уже удалено
                del user_resume_messages[user_id]

            # Выполняем удаление
            async with AsyncSessionLocal() as session:
                app_stmt = select(Application).where(Application.id == callback_data.application_id)
                app_result = await session.execute(app_stmt)
                application = app_result.scalar_one_or_none()

                if not application:
                    await query.message.edit_text("❌ Отклик не найден")
                    return

                # Удаляем файл если он существует
                if application.file_path and os.path.exists(application.file_path):
                    try:
                        os.remove(application.file_path)
                        print(f"Удален файл: {application.file_path}")
                    except Exception as e:
                        print(f"Ошибка удаления файла: {e}")

                # Используем soft delete вместо удаления из базы данных
                from datetime import datetime
                application.deleted_at = datetime.now()
                await session.commit()

                # Показываем уведомление об удалении
                await query.message.edit_text(
                    f"✅ Отклик от <b>{application.name}</b> успешно удален",
                    parse_mode="HTML"
                )

                # Ждем 1 секунду и возвращаемся к соответствующему меню
                import asyncio
                await asyncio.sleep(1)

                # Возвращаемся к соответствующему меню в зависимости от источника
                if callback_data.source == "unprocessed":
                    # Возвращаемся к списку необработанных откликов
                    async with AsyncSessionLocal() as session:
                        from sqlalchemy.orm import selectinload
                        stmt = select(Application).options(selectinload(Application.vacancy)).where(Application.is_processed == False, Application.deleted_at.is_(None)).order_by(desc(Application.created_at))
                        result = await session.execute(stmt)
                        unprocessed_applications = result.scalars().all()

                        if not unprocessed_applications:
                            await query.message.edit_text("✅ Все отклики обработаны!", parse_mode="HTML")
                            return

                        keyboard = InlineKeyboardMarkup(inline_keyboard=[])

                        for app in unprocessed_applications:
                            button_text = app.name
                            if app.phone:
                                button_text += f" - {app.phone}"
                            if app.email:
                                button_text += f" - {app.email}"

                            button_text = f"❌ {button_text}"

                            button = InlineKeyboardButton(
                                text=button_text,
                                callback_data=ApplicationCallback(application_id=app.id, source="unprocessed").pack()
                            )
                            keyboard.inline_keyboard.append([button])

                        text = f"❌ <b>Необработанные отклики ({len(unprocessed_applications)}):</b>\n\n"
                        text += "Выберите отклик для просмотра и обработки:"

                        await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

                elif callback_data.source == "vacancy":
                    # Возвращаемся к списку откликов вакансии
                    back_callback = BackCallback(to="applications", vacancy_id=application.vacancy_id)
                    await back_handler(query, back_callback, user)

                else:  # source == "recent" или другое
                    # Возвращаемся к списку вакансий
                    back_callback = BackCallback(to="vacancies")
                    await back_handler(query, back_callback, user)

        elif callback_data.action == "cancel":
            # Отменяем удаление - возвращаемся к детальному просмотру
            application_callback = ApplicationCallback(application_id=callback_data.application_id, source=callback_data.source)
            await application_details_handler(query, application_callback, user)

    @dp.callback_query(BackCallback.filter())
    async def back_handler(query: CallbackQuery, callback_data: BackCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав", show_alert=True)
            return

        # Удаляем сообщение с файлом резюме если оно есть
        user_id = query.from_user.id
        if user_id in user_resume_messages:
            try:
                await query.bot.delete_message(chat_id=query.message.chat.id, message_id=user_resume_messages[user_id])
            except Exception:
                pass  # Игнорируем ошибки если сообщение уже удалено
            del user_resume_messages[user_id]

        if callback_data.to == "vacancies":
            # Возвращаемся к списку вакансий
            async with AsyncSessionLocal() as session:
                stmt = select(Vacancy).order_by(desc(Vacancy.created_at))
                result = await session.execute(stmt)
                vacancies = result.scalars().all()

                if not vacancies:
                    await query.message.edit_text("Пока нет вакансий")
                    return

                keyboard = InlineKeyboardMarkup(inline_keyboard=[])

                for vacancy in vacancies:
                    count_stmt = select(Application).where(Application.vacancy_id == vacancy.id, Application.deleted_at.is_(None))
                    count_result = await session.execute(count_stmt)
                    applications_count = len(count_result.scalars().all())

                    button_text = f"{vacancy.title} ({applications_count} откликов)"
                    button = InlineKeyboardButton(
                        text=button_text,
                        callback_data=VacancyCallback(vacancy_id=vacancy.id).pack()
                    )
                    keyboard.inline_keyboard.append([button])

                await query.message.edit_text("📋 Выберите вакансию для просмотра откликов:", reply_markup=keyboard)

        elif callback_data.to == "applications":
            # Возвращаемся к списку откликов конкретной вакансии
            async with AsyncSessionLocal() as session:
                vacancy_stmt = select(Vacancy).where(Vacancy.id == callback_data.vacancy_id)
                vacancy_result = await session.execute(vacancy_stmt)
                vacancy = vacancy_result.scalar_one_or_none()

                if not vacancy:
                    await query.message.edit_text("Вакансия не найдена")
                    return

                apps_stmt = select(Application).where(Application.vacancy_id == callback_data.vacancy_id, Application.deleted_at.is_(None)).order_by(desc(Application.created_at))
                apps_result = await session.execute(apps_stmt)
                applications = apps_result.scalars().all()

                if not applications:
                    await query.message.edit_text(f"📋 Вакансия: {vacancy.title}\n\nОткликов пока нет")
                    return

                keyboard = InlineKeyboardMarkup(inline_keyboard=[])

                for app in applications:
                    button_text = app.name
                    if app.phone:
                        button_text += f" - {app.phone}"
                    if app.email:
                        button_text += f" - {app.email}"

                    status = "✅" if app.is_processed else "❌"
                    button_text = f"{status} {button_text}"

                    button = InlineKeyboardButton(
                        text=button_text,
                        callback_data=ApplicationCallback(application_id=app.id).pack()
                    )
                    keyboard.inline_keyboard.append([button])

                # Добавляем кнопку "Назад к вакансиям"
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад к вакансиям",
                    callback_data=BackCallback(to="vacancies").pack()
                )
                keyboard.inline_keyboard.append([back_button])

                text = f"📋 Вакансия: <b>{vacancy.title}</b>\n"
                text += f"📊 Всего откликов: <b>{len(applications)}</b>\n\n"
                text += "Выберите отклик для просмотра подробностей:"

                await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

        elif callback_data.to == "unprocessed":
            # Возвращаемся к списку необработанных откликов
            async with AsyncSessionLocal() as session:
                from sqlalchemy.orm import selectinload
                stmt = select(Application).options(selectinload(Application.vacancy)).where(Application.is_processed == False, Application.deleted_at.is_(None)).order_by(desc(Application.created_at))
                result = await session.execute(stmt)
                unprocessed_applications = result.scalars().all()

                if not unprocessed_applications:
                    await query.message.edit_text("✅ Все отклики обработаны!")
                    return

                keyboard = InlineKeyboardMarkup(inline_keyboard=[])

                for app in unprocessed_applications:
                    button_text = app.name
                    if app.phone:
                        button_text += f" - {app.phone}"
                    if app.email:
                        button_text += f" - {app.email}"

                    button_text = f"❌ {button_text}"

                    button = InlineKeyboardButton(
                        text=button_text,
                        callback_data=ApplicationCallback(application_id=app.id, source="unprocessed").pack()
                    )
                    keyboard.inline_keyboard.append([button])

                text = f"❌ <b>Необработанные отклики ({len(unprocessed_applications)}):</b>\n\n"
                text += "Выберите отклик для просмотра и обработки:"

                await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    async def do_export(message: Message, user: TelegramUser, filter_type: str = "all") -> None:
        status_msg = await message.answer("📊 Создаю Excel файл с откликами...")

        try:
            async with AsyncSessionLocal() as session:
                from sqlalchemy.orm import selectinload

                # Получаем все не удаленные отклики с вакансиями
                stmt = select(Application).options(selectinload(Application.vacancy)).where(
                    Application.deleted_at.is_(None)
                )

                # Добавляем фильтр по статусу обработки, если нужно
                if filter_type == "unprocessed":
                    stmt = stmt.where(Application.is_processed == False)

                stmt = stmt.order_by(desc(Application.created_at))
                result = await session.execute(stmt)
                applications = result.scalars().all()

                if not applications:
                    filter_text = "необработанных откликов" if filter_type == "unprocessed" else "откликов"
                    await status_msg.edit_text(f"📭 Нет {filter_text} для экспорта")
                    import asyncio
                    asyncio.create_task(delete_message_after_delay(status_msg, 2))
                    asyncio.create_task(delete_message_after_delay(message, 2))
                    return

                # Создаем Excel файл
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from datetime import datetime

                wb = Workbook()
                ws = wb.active
                ws.title = "Отклики на вакансии"

                # Заголовки
                headers = [
                    "ID", "Имя", "Email", "Телефон", "Вакансия",
                    "Статус", "Дата отклика", "Сообщение", "Описание обработки", "Файл", "Анализ резюме"
                ]

                # Стиль заголовков
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Данные
                for row, app in enumerate(applications, 2):
                    ws.cell(row=row, column=1, value=app.id)
                    ws.cell(row=row, column=2, value=app.name or "")
                    ws.cell(row=row, column=3, value=app.email or "")
                    ws.cell(row=row, column=4, value=app.phone or "")
                    ws.cell(row=row, column=5, value=app.vacancy.title if app.vacancy else "")
                    ws.cell(row=row, column=6, value="Обработан" if app.is_processed else "Не обработан")
                    ws.cell(row=row, column=7, value=app.created_at.strftime('%d.%m.%Y %H:%M') if app.created_at else "")

                    # Колонка "Сообщение" с переносом текста
                    message_cell = ws.cell(row=row, column=8, value=app.applicant_message or "")
                    message_cell.alignment = Alignment(wrap_text=True, vertical="top")

                    # Колонка "Описание обработки" с переносом текста
                    description_cell = ws.cell(row=row, column=9, value=app.processing_description or "")
                    description_cell.alignment = Alignment(wrap_text=True, vertical="top")

                    ws.cell(row=row, column=10, value=app.attachment_filename or "")

                    # Колонка "Анализ резюме" с очищенным от HTML текстом
                    summary_text = clean_html_tags(app.summary) if app.summary else ""
                    summary_cell = ws.cell(row=row, column=11, value=summary_text)
                    summary_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Сохраняем файл
                filename = f"отклики_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = f"exports/{filename}"

                # Создаем директорию если её нет
                import os
                os.makedirs("exports", exist_ok=True)

                wb.save(file_path)

                # Отправляем файл
                from aiogram.types import FSInputFile
                file = FSInputFile(file_path, filename=filename)

                export_type_text = "Только необработанные" if filter_type == "unprocessed" else "Все отклики"
                await status_msg.edit_text(f"✅ Excel файл готов!\nОтклики: {len(applications)}")
                await message.answer_document(
                    file,
                    caption=f"📊 Экспорт откликов\n\n"
                           f"Тип: {export_type_text}\n"
                           f"Всего откликов: {len(applications)}\n"
                           f"Создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )

                # Удаляем временный файл
                try:
                    os.remove(file_path)
                except:
                    pass

                # Удаляем сообщения через 2 секунды
                import asyncio
                asyncio.create_task(delete_message_after_delay(status_msg, 2))
                asyncio.create_task(delete_message_after_delay(message, 2))

        except Exception as e:
            await status_msg.edit_text(f"❌ Ошибка создания Excel: {str(e)}")
            import asyncio
            asyncio.create_task(delete_message_after_delay(status_msg, 2))
            asyncio.create_task(delete_message_after_delay(message, 2))

    @dp.message(Command("export"))
    @moderator_or_admin
    async def export_handler(message: Message, user: TelegramUser) -> None:
        # Создаем inline клавиатуру с выбором типа экспорта
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📊 Все отклики",
                    callback_data=ExportCallback(filter_type="all").pack()
                )
            ],
            [
                InlineKeyboardButton(
                    text="❌ Только необработанные",
                    callback_data=ExportCallback(filter_type="unprocessed").pack()
                )
            ]
        ])

        await message.answer(
            "📥 <b>Выберите тип экспорта:</b>",
            reply_markup=keyboard,
            parse_mode="HTML"
        )

    @dp.callback_query(ExportCallback.filter())
    async def export_callback_handler(query: CallbackQuery, callback_data: ExportCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав для экспорта", show_alert=True)
            return

        # Вызываем экспорт с выбранным фильтром
        await do_export(query.message, user, filter_type=callback_data.filter_type)

        # Удаляем сообщение с выбором после выполнения экспорта
        try:
            await query.message.delete()
        except Exception:
            pass

    @dp.callback_query(SummaryCallback.filter())
    async def summary_handler(query: CallbackQuery, callback_data: SummaryCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав", show_alert=True)
            return

        if callback_data.action == "show":
            # Показываем готовый summary
            async with AsyncSessionLocal() as session:
                # Получаем отклик
                app_stmt = select(Application).where(Application.id == callback_data.application_id)
                app_result = await session.execute(app_stmt)
                application = app_result.scalar_one_or_none()

                if not application:
                    await query.message.answer("❌ Отклик не найден")
                    return

                if application.summary:
                    # Отправляем готовый summary
                    summary_msg = f"📊 <b>Анализ резюме для {application.name}:</b>\n\n{application.summary}"
                    await query.message.answer(summary_msg, parse_mode="HTML")
                else:
                    await query.message.answer("❌ Анализ резюме не найден")

        elif callback_data.action == "generate":
            status_msg = await query.message.answer("🤖 Генерирую анализ резюме...")

            try:
                async with AsyncSessionLocal() as session:
                    # Получаем отклик
                    app_stmt = select(Application).where(Application.id == callback_data.application_id)
                    app_result = await session.execute(app_stmt)
                    application = app_result.scalar_one_or_none()

                    if not application:
                        await status_msg.edit_text("❌ Отклик не найден")
                        return

                    # Получаем вакансию
                    vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
                    vacancy_result = await session.execute(vacancy_stmt)
                    vacancy = vacancy_result.scalar_one_or_none()

                    # Создаем сервис для генерации summary
                    summary_service = ResumeSummaryService()

                    # Генерируем summary
                    summary = await summary_service.generate_summary_for_application(application, vacancy)

                    if summary:
                        # Сохраняем summary в базу данных
                        application.summary = summary
                        await session.commit()

                        await status_msg.edit_text("✅ Анализ резюме сгенерирован!")

                        # Отправляем сгенерированный summary с кнопкой для генерации вопросов
                        summary_msg = f"🤖 <b>Анализ резюме для {application.name}:</b>\n\n{summary}"
                        questions_keyboard = InlineKeyboardMarkup(inline_keyboard=[
                            [InlineKeyboardButton(
                                text="❓ Сгенерировать вопросы для собеседования",
                                callback_data=QuestionsCallback(application_id=application.id, action="generate").pack()
                            )]
                        ])
                        await query.message.answer(summary_msg, parse_mode="HTML", reply_markup=questions_keyboard)

                        # Удаляем статусное сообщение через 2 секунды
                        import asyncio
                        asyncio.create_task(delete_message_after_delay(status_msg, 2))

                    else:
                        await status_msg.edit_text("❌ Не удалось сгенерировать анализ резюме")
                        import asyncio
                        asyncio.create_task(delete_message_after_delay(status_msg, 3))

            except Exception as e:
                await status_msg.edit_text(f"❌ Ошибка при генерации анализа: {str(e)}")
                import asyncio
                asyncio.create_task(delete_message_after_delay(status_msg, 3))

    @dp.callback_query(QuestionsCallback.filter())
    async def questions_handler(query: CallbackQuery, callback_data: QuestionsCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав", show_alert=True)
            return

        if callback_data.action == "generate":
            status_msg = await query.message.answer("🤖 Генерирую вопросы для собеседования...")

            try:
                async with AsyncSessionLocal() as session:
                    # Получаем отклик
                    app_stmt = select(Application).where(Application.id == callback_data.application_id)
                    app_result = await session.execute(app_stmt)
                    application = app_result.scalar_one_or_none()

                    if not application:
                        await status_msg.edit_text("❌ Отклик не найден")
                        return

                    # Проверяем наличие файла резюме
                    if not application.file_path:
                        await status_msg.edit_text("❌ Файл резюме не найден")
                        return

                    # Получаем вакансию
                    vacancy_stmt = select(Vacancy).where(Vacancy.id == application.vacancy_id)
                    vacancy_result = await session.execute(vacancy_stmt)
                    vacancy = vacancy_result.scalar_one_or_none()

                    # Извлекаем текст из резюме
                    from shared.services.document_extractor import DocumentTextExtractor
                    from shared.services.gemini_service import GeminiService

                    extractor = DocumentTextExtractor()
                    resume_text = extractor.extract_text_from_file(application.file_path)

                    if not resume_text:
                        await status_msg.edit_text("❌ Не удалось извлечь текст из резюме")
                        return

                    # Генерируем вопросы через Gemini
                    gemini_service = GeminiService()
                    vacancy_title = vacancy.title if vacancy else ""
                    questions = gemini_service.generate_interview_questions(resume_text, vacancy_title)

                    if questions:
                        await status_msg.edit_text("✅ Вопросы сгенерированы!")

                        # Отправляем вопросы
                        await query.message.answer(questions, parse_mode="HTML")

                        # Удаляем статусное сообщение через 2 секунды
                        import asyncio
                        asyncio.create_task(delete_message_after_delay(status_msg, 2))
                    else:
                        await status_msg.edit_text("❌ Не удалось сгенерировать вопросы")
                        import asyncio
                        asyncio.create_task(delete_message_after_delay(status_msg, 3))

            except Exception as e:
                await status_msg.edit_text(f"❌ Ошибка при генерации вопросов: {str(e)}")
                import asyncio
                asyncio.create_task(delete_message_after_delay(status_msg, 3))

    @dp.callback_query(ResumeCallback.filter())
    async def resume_handler(query: CallbackQuery, callback_data: ResumeCallback, user: TelegramUser) -> None:
        await query.answer()

        if not user.has_permission('view_applications'):
            await query.answer("❌ У вас нет прав", show_alert=True)
            return

        if callback_data.action == "download":
            async with AsyncSessionLocal() as session:
                # Получаем отклик
                app_stmt = select(Application).where(Application.id == callback_data.application_id)
                app_result = await session.execute(app_stmt)
                application = app_result.scalar_one_or_none()

                if not application:
                    await query.message.answer("❌ Отклик не найден")
                    return

                # Отправляем файл если он есть
                user_id = query.from_user.id
                if application.file_path and os.path.exists(application.file_path):
                    try:
                        from aiogram.types import FSInputFile
                        file = FSInputFile(application.file_path, filename=application.attachment_filename)
                        file_msg = await query.message.answer_document(file, caption=f"📄 Резюме от {application.name}")
                        # Сохраняем ID сообщения с файлом
                        user_resume_messages[user_id] = file_msg.message_id
                    except Exception as e:
                        error_msg = await query.message.answer(f"❌ Ошибка при отправке файла: {str(e)}")
                        user_resume_messages[user_id] = error_msg.message_id
                elif application.file_url:
                    url_msg = await query.message.answer(f"📄 Файл резюме доступен по ссылке: {application.file_url}")
                    user_resume_messages[user_id] = url_msg.message_id
                else:
                    await query.message.answer("❌ Файл резюме не найден")

    @dp.message(Command("accounts"))
    @admin_only
    async def accounts_handler(message: Message, user: TelegramUser) -> None:
        """Показывает список всех Gmail аккаунтов"""
        async with AsyncSessionLocal() as session:
            from shared.models.gmail_account import GmailAccount

            # Получаем все аккаунты из БД
            stmt = select(GmailAccount).order_by(GmailAccount.name)
            result = await session.execute(stmt)
            accounts = result.scalars().all()

            if not accounts:
                await message.answer("📭 Нет настроенных Gmail аккаунтов\n\n"
                                   "Используйте команду /add_account для добавления")
                return

            # Создаем клавиатуру с аккаунтами
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for account in accounts:
                status_emoji = "✅" if account.enabled else "❌"
                button_text = f"{status_emoji} {account.name}"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=AccountCallback(account_id=account.account_id).pack()
                )
                keyboard.inline_keyboard.append([button])

            text = "📧 <b>Gmail аккаунты</b>\n\n"
            text += f"Всего аккаунтов: <b>{len(accounts)}</b>\n"
            enabled_count = sum(1 for acc in accounts if acc.enabled)
            text += f"Активных: <b>{enabled_count}</b>\n\n"
            text += "Выберите аккаунт для просмотра деталей:"

            await message.answer(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(AccountCallback.filter())
    async def account_details_handler(query: CallbackQuery, callback_data: AccountCallback, user: TelegramUser) -> None:
        """Показывает детали конкретного аккаунта"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            from shared.models.gmail_account import GmailAccount
            from sqlalchemy.orm import selectinload

            # Получаем аккаунт из БД по account_id (строка)
            stmt = select(GmailAccount).options(selectinload(GmailAccount.user)).where(
                GmailAccount.account_id == callback_data.account_id
            )
            result = await session.execute(stmt)
            account = result.scalar_one_or_none()

            if not account:
                await query.message.edit_text("❌ Аккаунт не найден")
                return

            # Формируем детальную информацию
            status_emoji = "✅" if account.enabled else "❌"
            status_text = "Активен" if account.enabled else "Отключен"

            text = f"📧 <b>{account.name}</b>\n\n"
            text += f"🆔 <b>ID:</b> <code>{account.id}</code>\n"
            text += f"🏷️ <b>Статус:</b> {status_emoji} {status_text}\n"

            # Показываем привязанного пользователя
            if account.user:
                role_emoji = "👑" if account.user.is_admin else ("👨‍💼" if account.user.is_moderator else "👤")
                text += f"👤 <b>Привязан к:</b> {role_emoji} {account.user.first_name or account.user.username or f'ID: {account.user.telegram_id}'}\n"
            else:
                text += f"👤 <b>Привязка:</b> Не привязан\n"

            # Создаем кнопки управления
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            if account.enabled:
                toggle_button = InlineKeyboardButton(
                    text="❌ Отключить аккаунт",
                    callback_data=AccountToggleCallback(account_id=account.account_id, action="disable").pack()
                )
            else:
                toggle_button = InlineKeyboardButton(
                    text="✅ Включить аккаунт",
                    callback_data=AccountToggleCallback(account_id=account.account_id, action="enable").pack()
                )

            keyboard.inline_keyboard.append([toggle_button])

            # Кнопка "Привязать" или "Отвязать"
            if account.user:
                # Показываем кнопку отвязки
                unlink_button = InlineKeyboardButton(
                    text="🔓 Отвязать от пользователя",
                    callback_data=AccountLinkCallback(account_id=account.account_id, action="unlink").pack()
                )
                keyboard.inline_keyboard.append([unlink_button])
            else:
                # Показываем кнопку привязки
                link_button = InlineKeyboardButton(
                    text="👤 Привязать к пользователю",
                    callback_data=AccountLinkCallback(account_id=account.account_id, action="show_users").pack()
                )
                keyboard.inline_keyboard.append([link_button])

            # Кнопка "Авторизация"
            auth_button = InlineKeyboardButton(
                text="🔐 Переавторизовать",
                callback_data=AccountAuthCallback(account_id=account.account_id).pack()
            )
            keyboard.inline_keyboard.append([auth_button])

            # Кнопка "Удалить аккаунт"
            delete_button = InlineKeyboardButton(
                text="🗑️ Удалить аккаунт",
                callback_data=AccountDeleteCallback(account_id=account.account_id, action="confirm").pack()
            )
            keyboard.inline_keyboard.append([delete_button])

            # Кнопка "Назад"
            back_button = InlineKeyboardButton(
                text="⬅️ Назад к списку",
                callback_data="back_to_accounts"
            )
            keyboard.inline_keyboard.append([back_button])

            await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(AccountLinkCallback.filter())
    async def account_link_handler(query: CallbackQuery, callback_data: AccountLinkCallback, user: TelegramUser) -> None:
        """Обработчик привязки аккаунта к пользователю"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        if callback_data.action == "show_users":
            # Показываем список пользователей для привязки
            async with AsyncSessionLocal() as session:
                # Выбираем всех пользователей
                stmt = select(TelegramUser).order_by(TelegramUser.role.desc(), TelegramUser.first_name)
                result = await session.execute(stmt)
                users = result.scalars().all()

                if not users:
                    await query.message.edit_text("❌ Нет пользователей для привязки")
                    return

                keyboard = InlineKeyboardMarkup(inline_keyboard=[])
                text = f"👥 <b>Выберите пользователя для привязки аккаунта:</b>\n\n"

                for u in users:
                    role_emoji = "👑" if u.is_admin else ("👨‍💼" if u.is_moderator else "👤")
                    role_text = u.role.value.upper()
                    user_text = f"{role_emoji} {u.first_name or u.username or f'ID: {u.telegram_id}'} ({role_text})"
                    button = InlineKeyboardButton(
                        text=user_text,
                        callback_data=AccountLinkCallback(
                            account_id=callback_data.account_id,
                            action="link",
                            user_id=u.id
                        ).pack()
                    )
                    keyboard.inline_keyboard.append([button])

                # Кнопка назад
                back_button = InlineKeyboardButton(
                    text="⬅️ Назад",
                    callback_data=AccountCallback(account_id=callback_data.account_id).pack()
                )
                keyboard.inline_keyboard.append([back_button])

                await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

        elif callback_data.action == "unlink":
            # Отвязываем аккаунт от пользователя
            async with AsyncSessionLocal() as session:
                from shared.models.gmail_account import GmailAccount

                # Ищем по account_id (строка)
                stmt = select(GmailAccount).where(GmailAccount.account_id == callback_data.account_id)
                result = await session.execute(stmt)
                gmail_account = result.scalar_one_or_none()
                if not gmail_account:
                    await query.message.edit_text("❌ Аккаунт не найден")
                    return

                # Отвязываем
                gmail_account.user_id = None
                await session.commit()

                await query.answer("✅ Аккаунт отвязан от пользователя", show_alert=True)

                # Возвращаемся к деталям аккаунта
                # Эмулируем callback для показа обновленных деталей
                callback_obj = AccountCallback(account_id=callback_data.account_id)
                await account_details_handler(query, callback_obj, user)

        elif callback_data.action == "link":
            # Привязываем аккаунт к пользователю
            async with AsyncSessionLocal() as session:
                from shared.models.gmail_account import GmailAccount

                # Ищем аккаунт в БД по account_id (строка)
                stmt = select(GmailAccount).where(GmailAccount.account_id == callback_data.account_id)
                result = await session.execute(stmt)
                gmail_account = result.scalar_one_or_none()

                if not gmail_account:
                    await query.message.edit_text("❌ Аккаунт не найден")
                    return

                # Привязываем к пользователю
                gmail_account.user_id = callback_data.user_id
                await session.commit()

                # Получаем пользователя для отображения
                linked_user = await session.get(TelegramUser, callback_data.user_id)
                user_name = linked_user.first_name or linked_user.username or f"ID: {linked_user.telegram_id}"

                await query.message.edit_text(
                    f"✅ Аккаунт <b>{gmail_account.name}</b> привязан к пользователю <b>{user_name}</b>",
                    parse_mode="HTML"
                )

                # Возвращаемся к деталям аккаунта через 2 секунды
                import asyncio
                await asyncio.sleep(2)
                callback_obj = AccountCallback(account_id=callback_data.account_id)
                await account_details_handler(query, callback_obj, user)

    @dp.callback_query(AccountAuthCallback.filter())
    async def account_auth_handler(query: CallbackQuery, callback_data: AccountAuthCallback, user: TelegramUser) -> None:
        """Обработчик авторизации аккаунта"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        from bot.gmail_account_manager import GmailAccountManager

        # Генерируем URL авторизации
        success, result, _ = GmailAccountManager.generate_auth_url()

        if not success:
            await query.answer(result, show_alert=True)
            return

        auth_url = result

        # Сохраняем состояние ожидания кода для этого пользователя
        user_auth_states[query.from_user.id] = True

        # Отправляем URL в личку пользователю
        await query.message.answer(
            f"🔐 <b>Переавторизация аккаунта</b>\n\n"
            f"1️⃣ Перейдите по ссылке:\n{auth_url}\n\n"
            f"2️⃣ Авторизуйтесь в Google аккаунте и разрешите доступ\n\n"
            f"3️⃣ После авторизации браузер покажет ошибку — <b>это нормально!</b>\n\n"
            f"4️⃣ Скопируйте <b>весь URL</b> из адресной строки браузера и отправьте мне\n\n"
            f"💡 URL будет выглядеть примерно так:\n"
            f"<code>http://localhost/?code=4/0Adeu5BW...&scope=...</code>",
            parse_mode="HTML",
            disable_web_page_preview=True
        )

        await query.answer("✅ Ссылка отправлена вам в сообщении", show_alert=True)

    @dp.callback_query(AccountDeleteCallback.filter())
    async def account_delete_handler(query: CallbackQuery, callback_data: AccountDeleteCallback, user: TelegramUser) -> None:
        """Обработчик удаления аккаунта"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        if callback_data.action == "confirm":
            # Показываем подтверждение удаления
            async with AsyncSessionLocal() as session:
                from shared.models.gmail_account import GmailAccount

                # Ищем по account_id (строка)
                stmt = select(GmailAccount).where(GmailAccount.account_id == callback_data.account_id)
                result = await session.execute(stmt)
                account = result.scalar_one_or_none()
                if not account:
                    await query.message.edit_text("❌ Аккаунт не найден")
                    return

                keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Да, удалить",
                            callback_data=AccountDeleteCallback(account_id=callback_data.account_id, action="execute").pack()
                        ),
                        InlineKeyboardButton(
                            text="❌ Отмена",
                            callback_data=AccountCallback(account_id=callback_data.account_id).pack()
                        )
                    ]
                ])

                text = f"⚠️ <b>Подтверждение удаления</b>\n\n"
                text += f"Вы действительно хотите удалить аккаунт?\n\n"
                text += f"📧 <b>{account.name}</b>\n"
                text += f"🆔 <code>{account.id}</code>\n\n"
                text += f"❗ <b>Внимание:</b> Все вакансии и отклики, связанные с этим аккаунтом, останутся в базе, но потеряют связь с аккаунтом."

                await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

        elif callback_data.action == "execute":
            # Удаляем аккаунт
            async with AsyncSessionLocal() as session:
                from shared.models.gmail_account import GmailAccount

                # Ищем по account_id (строка)
                stmt = select(GmailAccount).where(GmailAccount.account_id == callback_data.account_id)
                result = await session.execute(stmt)
                account = result.scalar_one_or_none()
                if not account:
                    await query.message.edit_text("❌ Аккаунт не найден")
                    return

                account_name = account.name

                # Удаляем из БД
                await session.delete(account)
                await session.commit()

                await query.message.edit_text(
                    f"✅ Аккаунт <b>{account_name}</b> успешно удален",
                    parse_mode="HTML"
                )

                # Возвращаемся к списку через 2 секунды
                import asyncio
                await asyncio.sleep(2)

                # Эмулируем нажатие кнопки "Назад к списку"
                from aiogram.types import Message as TgMessage
                fake_message = TgMessage(
                    message_id=query.message.message_id,
                    date=query.message.date,
                    chat=query.message.chat,
                    from_user=query.from_user
                )
                await accounts_handler(fake_message, user)

    @dp.callback_query(AccountToggleCallback.filter())
    async def account_toggle_handler(query: CallbackQuery, callback_data: AccountToggleCallback, user: TelegramUser) -> None:
        """Включает или отключает аккаунт"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            from shared.models.gmail_account import GmailAccount

            # Получаем аккаунт из БД по account_id (строка)
            stmt = select(GmailAccount).where(GmailAccount.account_id == callback_data.account_id)
            result = await session.execute(stmt)
            account = result.scalar_one_or_none()

            if not account:
                await query.message.answer("❌ Аккаунт не найден")
                return

            # Обновляем статус
            if callback_data.action == "enable":
                account.enabled = True
                status_msg = f"✅ Аккаунт <b>{account.name}</b> включен"
            else:
                account.enabled = False
                status_msg = f"❌ Аккаунт <b>{account.name}</b> отключен"

            await session.commit()

            # Показываем уведомление
            notification = await query.message.answer(status_msg, parse_mode="HTML")

            # Обновляем детали аккаунта
            callback_obj = AccountCallback(account_id=callback_data.account_id)
            await account_details_handler(query, callback_obj, user)

            # Удаляем уведомление через 2 секунды
            import asyncio
            asyncio.create_task(delete_message_after_delay(notification, 2))

    @dp.callback_query(lambda c: c.data == "back_to_accounts")
    async def back_to_accounts_handler(query: CallbackQuery, user: TelegramUser) -> None:
        """Возвращается к списку аккаунтов"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Только для администраторов", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            from shared.models.gmail_account import GmailAccount

            # Получаем все аккаунты из БД
            stmt = select(GmailAccount).order_by(GmailAccount.name)
            result = await session.execute(stmt)
            accounts = result.scalars().all()

            # Создаем клавиатуру с аккаунтами
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for account in accounts:
                status_emoji = "✅" if account.enabled else "❌"
                button_text = f"{status_emoji} {account.name}"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=AccountCallback(account_id=account.account_id).pack()
                )
                keyboard.inline_keyboard.append([button])

            text = "📧 <b>Gmail аккаунты</b>\n\n"
            text += f"Всего аккаунтов: <b>{len(accounts)}</b>\n"
            enabled_count = sum(1 for acc in accounts if acc.enabled)
            text += f"Активных: <b>{enabled_count}</b>\n\n"
            text += "Выберите аккаунт для просмотра деталей:"

            await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    # Словарь для хранения состояния авторизации пользователей
    user_auth_states = {}

    @dp.message(Command("add_account", "add_gmail"))
    @admin_only
    async def add_account_command_handler(message: Message, user: TelegramUser) -> None:
        """Добавляет новый Gmail аккаунт через OAuth"""
        from bot.gmail_account_manager import GmailAccountManager

        try:
            # Генерируем OAuth URL
            success, auth_url, flow_data = GmailAccountManager.generate_auth_url()

            if not success:
                await message.answer(auth_url, parse_mode="HTML")
                return

            # Сохраняем состояние ожидания кода для этого пользователя
            user_auth_states[message.from_user.id] = True

            auth_msg = (
                "🔐 <b>Добавление нового Gmail аккаунта</b>\n\n"
                "1️⃣ Перейдите по ссылке ниже для авторизации:\n"
                f"<a href='{auth_url}'>Открыть страницу авторизации Google</a>\n\n"
                "2️⃣ Выберите Gmail аккаунт и разрешите доступ\n\n"
                "3️⃣ После авторизации браузер покажет ошибку — <b>это нормально!</b>\n\n"
                "4️⃣ Скопируйте <b>весь URL</b> из адресной строки браузера и отправьте мне\n\n"
                "💡 URL будет выглядеть примерно так:\n"
                "<code>http://localhost/?code=4/0Adeu5BW...&scope=...</code>"
            )

            await message.answer(auth_msg, parse_mode="HTML", disable_web_page_preview=True)

        except Exception as e:
            error_text = (
                f"❌ <b>Ошибка генерации ссылки</b>\n\n"
                f"<code>{str(e)}</code>\n\n"
                f"💡 Убедитесь, что credentials.json находится в gmail_tokens/"
            )
            await message.answer(error_text, parse_mode="HTML")

    @dp.message(lambda message: message.from_user.id in user_auth_states and user_auth_states.get(message.from_user.id) and message.text and not message.text.startswith('/'))
    async def handle_auth_code(message: Message, user: TelegramUser) -> None:
        """Обрабатывает код авторизации от пользователя"""
        from bot.gmail_account_manager import GmailAccountManager

        if not user.is_admin:
            await message.answer("❌ Только для администраторов")
            return

        auth_input = message.text.strip()

        # Извлекаем код из URL если пользователь отправил весь URL
        auth_code = auth_input
        if 'code=' in auth_input:
            from urllib.parse import urlparse, parse_qs
            try:
                parsed = urlparse(auth_input)
                code_params = parse_qs(parsed.query).get('code', [])
                if code_params:
                    auth_code = code_params[0]
            except Exception:
                pass

        # Проверяем что это похоже на код авторизации
        if not auth_code.startswith('4/'):
            await message.answer(
                "⚠️ Это не похоже на код авторизации.\n\n"
                "Скопируйте весь URL из адресной строки браузера.\n"
                "URL должен содержать <code>code=4/...</code>\n\n"
                "Попробуйте еще раз или отправьте /cancel для отмены.",
                parse_mode="HTML"
            )
            return

        status_msg = await message.answer("⏳ Проверяю код авторизации...")

        try:
            # Завершаем авторизацию с кодом
            success, msg, account_data = await GmailAccountManager.complete_auth_with_code(auth_code)

            # Убираем состояние ожидания кода
            del user_auth_states[message.from_user.id]

            if success:
                # Успешно добавлен
                final_text = (
                    "✅ <b>Новый аккаунт успешно добавлен!</b>\n\n"
                    f"📧 <b>Email:</b> <code>{account_data['name']}</code>\n"
                    f"🆔 <b>ID:</b> <code>{account_data['id']}</code>\n"
                    f"🏷️ <b>Статус:</b> ❌ Отключен\n\n"
                    "💡 <b>Следующие шаги:</b>\n"
                    "1. Используйте /accounts\n"
                    "2. Выберите этот аккаунт\n"
                    "3. Нажмите \"✅ Включить аккаунт\""
                )
                await status_msg.edit_text(final_text, parse_mode="HTML")
            else:
                # Ошибка
                await status_msg.edit_text(msg, parse_mode="HTML")

            # Удаляем сообщение с кодом (для безопасности)
            import asyncio
            asyncio.create_task(delete_message_after_delay(message, 1))

        except Exception as e:
            # Убираем состояние ожидания кода
            if message.from_user.id in user_auth_states:
                del user_auth_states[message.from_user.id]

            error_text = (
                f"❌ <b>Ошибка авторизации</b>\n\n"
                f"<code>{str(e)}</code>\n\n"
                f"💡 Попробуйте еще раз: /add_account"
            )
            await status_msg.edit_text(error_text, parse_mode="HTML")

            # Удаляем сообщение с кодом
            import asyncio
            asyncio.create_task(delete_message_after_delay(message, 1))

    @dp.message(Command("users"))
    @admin_only
    async def users_handler(message: Message, user: TelegramUser) -> None:
        """Показывает список всех пользователей бота"""
        async with AsyncSessionLocal() as session:
            stmt = select(TelegramUser).order_by(TelegramUser.created_at.desc())
            result = await session.execute(stmt)
            users = result.scalars().all()

            if not users:
                await message.answer("📭 Нет пользователей")
                return

            # Создаем клавиатуру с пользователями
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for u in users:
                role_emoji = {
                    RoleEnum.USER: "👤",
                    RoleEnum.MODERATOR: "👨‍💼",
                    RoleEnum.ADMIN: "👑"
                }.get(u.role, "👤")

                button_text = f"{role_emoji} {u.first_name or 'Unknown'}"
                if u.username:
                    button_text += f" (@{u.username})"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=UserCallback(user_id=u.id).pack()
                )
                keyboard.inline_keyboard.append([button])

            text = "👥 <b>Пользователи бота</b>\n\n"
            text += f"Всего: <b>{len(users)}</b>\n\n"
            text += "Выберите пользователя для управления:"

            await message.answer(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(UserCallback.filter())
    async def user_details_handler(query: CallbackQuery, callback_data: UserCallback, user: TelegramUser) -> None:
        """Показывает детали пользователя"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Недостаточно прав", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            stmt = select(TelegramUser).where(TelegramUser.id == callback_data.user_id)
            result = await session.execute(stmt)
            selected_user = result.scalar_one_or_none()

            if not selected_user:
                await query.message.edit_text("❌ Пользователь не найден")
                return

            role_emoji = {
                RoleEnum.USER: "👤",
                RoleEnum.MODERATOR: "👨‍💼",
                RoleEnum.ADMIN: "👑"
            }.get(selected_user.role, "👤")

            role_name = {
                RoleEnum.USER: "Пользователь",
                RoleEnum.MODERATOR: "Модератор",
                RoleEnum.ADMIN: "Администратор"
            }.get(selected_user.role, "Пользователь")

            text = f"{role_emoji} <b>{selected_user.first_name or 'Unknown'}</b>\n\n"
            if selected_user.username:
                text += f"🆔 @{selected_user.username}\n"
            text += f"🔢 Telegram ID: <code>{selected_user.telegram_id}</code>\n"
            text += f"🏷️ Роль: <b>{role_name}</b>\n"
            text += f"📅 Регистрация: {selected_user.created_at.strftime('%d.%m.%Y %H:%M')}\n\n"

            # Кнопки для изменения роли
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            if selected_user.role != RoleEnum.ADMIN:
                admin_button = InlineKeyboardButton(
                    text="👑 Сделать администратором",
                    callback_data=UserRoleCallback(user_id=selected_user.id, role="admin").pack()
                )
                keyboard.inline_keyboard.append([admin_button])

            if selected_user.role != RoleEnum.MODERATOR:
                mod_button = InlineKeyboardButton(
                    text="👨‍💼 Сделать модератором",
                    callback_data=UserRoleCallback(user_id=selected_user.id, role="moderator").pack()
                )
                keyboard.inline_keyboard.append([mod_button])

            if selected_user.role != RoleEnum.USER:
                user_button = InlineKeyboardButton(
                    text="👤 Сделать пользователем",
                    callback_data=UserRoleCallback(user_id=selected_user.id, role="user").pack()
                )
                keyboard.inline_keyboard.append([user_button])

            # Кнопка "Назад"
            back_button = InlineKeyboardButton(
                text="⬅️ Назад к списку",
                callback_data="back_to_users"
            )
            keyboard.inline_keyboard.append([back_button])

            await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.callback_query(UserRoleCallback.filter())
    async def user_role_change_handler(query: CallbackQuery, callback_data: UserRoleCallback, user: TelegramUser) -> None:
        """Изменяет роль пользователя"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Недостаточно прав", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            stmt = select(TelegramUser).where(TelegramUser.id == callback_data.user_id)
            result = await session.execute(stmt)
            selected_user = result.scalar_one_or_none()

            if not selected_user:
                await query.message.answer("❌ Пользователь не найден")
                return

            # Запрещаем себя понижать, если это единственный админ
            if user.id == selected_user.id and user.is_admin:
                # Проверяем, есть ли другие админы
                admin_count_stmt = select(TelegramUser).where(TelegramUser.role == RoleEnum.ADMIN)
                admin_result = await session.execute(admin_count_stmt)
                admins = admin_result.scalars().all()

                if len(admins) <= 1 and callback_data.role != "admin":
                    await query.answer("❌ Нельзя изменить роль последнего администратора", show_alert=True)
                    return

            # Изменяем роль
            old_role = selected_user.role
            new_role = RoleEnum[callback_data.role.upper()]
            selected_user.role = new_role
            await session.commit()

            role_name = {
                RoleEnum.USER: "Пользователь",
                RoleEnum.MODERATOR: "Модератор",
                RoleEnum.ADMIN: "Администратор"
            }.get(new_role, "Пользователь")

            # Показываем уведомление
            notification = await query.message.answer(
                f"✅ Роль изменена на <b>{role_name}</b>",
                parse_mode="HTML"
            )

            # Обновляем детали пользователя
            user_callback = UserCallback(user_id=selected_user.id)
            await user_details_handler(query, user_callback, user)

            # Удаляем уведомление через 2 секунды
            import asyncio
            asyncio.create_task(delete_message_after_delay(notification, 2))

    @dp.callback_query(lambda c: c.data == "back_to_users")
    async def back_to_users_handler(query: CallbackQuery, user: TelegramUser) -> None:
        """Возвращается к списку пользователей"""
        await query.answer()

        if not user.is_admin:
            await query.answer("❌ Недостаточно прав", show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            stmt = select(TelegramUser).order_by(TelegramUser.created_at.desc())
            result = await session.execute(stmt)
            users = result.scalars().all()

            # Создаем клавиатуру с пользователями
            keyboard = InlineKeyboardMarkup(inline_keyboard=[])

            for u in users:
                role_emoji = {
                    RoleEnum.USER: "👤",
                    RoleEnum.MODERATOR: "👨‍💼",
                    RoleEnum.ADMIN: "👑"
                }.get(u.role, "👤")

                button_text = f"{role_emoji} {u.first_name or 'Unknown'}"
                if u.username:
                    button_text += f" (@{u.username})"

                button = InlineKeyboardButton(
                    text=button_text,
                    callback_data=UserCallback(user_id=u.id).pack()
                )
                keyboard.inline_keyboard.append([button])

            text = "👥 <b>Пользователи бота</b>\n\n"
            text += f"Всего: <b>{len(users)}</b>\n\n"
            text += "Выберите пользователя для управления:"

            await query.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")

    @dp.message(Command("cancel"))
    async def cancel_handler(message: Message) -> None:
        """Отменяет текущую операцию"""
        user_id = message.from_user.id

        if user_id in user_auth_states:
            del user_auth_states[user_id]
            await message.answer("✅ Операция отменена")
        elif user_id in user_description_states:
            del user_description_states[user_id]
            await message.answer("✅ Ввод описания отменён")
        else:
            await message.answer("❌ Нет активных операций для отмены")

    # Обработчики текстовых команд с эмодзи
    @dp.message(lambda message: message.text == "🏠 Главная")
    async def text_start_handler(message: Message, user: TelegramUser) -> None:
        await command_start_handler(message, user)

    @dp.message(lambda message: message.text == "📋 Последние")
    async def text_recent_handler(message: Message, user: TelegramUser) -> None:
        await recent_handler(message, user)

    @dp.message(lambda message: message.text == "⏳ Необработанные")
    async def text_unprocessed_handler(message: Message, user: TelegramUser) -> None:
        await unprocessed_handler(message, user)

    @dp.message(lambda message: message.text == "🔄 Парсинг")
    async def text_parse_handler(message: Message, user: TelegramUser) -> None:
        await parse_handler(message, user)

    @dp.message(lambda message: message.text == "📊 Статистика")
    async def text_stats_handler(message: Message, user: TelegramUser) -> None:
        await stats_handler(message, user)

    @dp.message(lambda message: message.text == "📧 Аккаунты")
    async def text_accounts_handler(message: Message, user: TelegramUser) -> None:
        await accounts_handler(message, user)

    @dp.message(lambda message: message.text == "➕ Добавить аккаунт")
    async def text_add_account_handler(message: Message, user: TelegramUser) -> None:
        await add_account_command_handler(message, user)

    @dp.message(lambda message: message.text == "📥 Экспорт")
    async def text_export_handler(message: Message, user: TelegramUser) -> None:
        if not user.has_permission('view_applications'):
            await message.answer("❌ У вас нет прав для экспорта")
            return

        # Создаем inline клавиатуру с выбором типа экспорта
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📊 Все отклики",
                    callback_data=ExportCallback(filter_type="all").pack()
                )
            ],
            [
                InlineKeyboardButton(
                    text="❌ Только необработанные",
                    callback_data=ExportCallback(filter_type="unprocessed").pack()
                )
            ]
        ])

        await message.answer(
            "📥 <b>Выберите тип экспорта:</b>",
            reply_markup=keyboard,
            parse_mode="HTML"
        )

    @dp.message(lambda message: message.text == "👥 Пользователи")
    async def text_users_handler(message: Message, user: TelegramUser) -> None:
        await users_handler(message, user)
